import openpyxl

from tools.file_base_path import BASE_PATH

excel = openpyxl.load_workbook(BASE_PATH + 'login_data.xlsx')
# 激活当前表
sheet = excel.active
# print(list(sheet.values))
# print(sheet.max_colum)

# def MaxRow(sheet):
#     maxrow = 0
#     maxcol = 0
#     for x in sheet.max_row:
#         for y in sheet.max_col:
#             if y == None:
#                 maxcol = y
#                 continue
#         if x == None:
#             maxrow = int(x)
#             return maxrow
#
#
#
list_login_xlsx = []
# max_row_num = MaxRow(sheet)
# print(max_row_num)
# print(list(sheet.values))
for row in sheet.iter_rows(min_row=2, min_col=1, max_row=2, max_col=2):
    row_data = [cell.value for cell in row]
    list_login_xlsx.append(tuple(row_data))

# print(list_login_xlsx)